"""
Batch processor for inserting data from multiple XML invoices into a
hierarchical mapping Excel file using JSON arrays for repeated fields.

This script searches for all `.xml` files in a subdirectory named
`XML` located in the same directory as this script. For each XML file
found, it extracts the values defined by the ID path mapping (from the
official representation) and appends a new row to the Excel file.
If a tag appears multiple times in the XML, its cell will contain a
JSON array of the values (e.g. `["1", "2", "3"]`).

Usage::

    python hierarchical_xml_to_excel_by_id_JSON.py

The script assumes that the following files are present in the same
directory as this script:

* ``RappresentazioneTabellareFattOrdinaria (1).xlsx`` – the official
  representation to map IDs to tag paths.
* ``hierarchical_mapping_fixed.xlsx`` – the hierarchical mapping
  produced previously.
* ``hierarchical_mapping_populated_json.xlsx`` – the Excel file into
  which values will be appended.
"""

import os
import json
from typing import Dict, List
import xml.etree.ElementTree as ET
from collections import defaultdict
import re

try:
    import openpyxl  # type: ignore
    import pandas as pd  # type: ignore
except ImportError:
    raise SystemExit("This script requires openpyxl and pandas. Please install them via pip.")


def build_id_to_tag_map(representation_xlsx: str) -> Dict[str, str]:
    """Build a mapping from ID code to tag name by reading the representation Excel."""
    df = pd.read_excel(representation_xlsx, header=None)
    header_idx = df.index[
        df.iloc[:, 0].astype(str).str.contains('ID e Nome', na=False)
    ].tolist()[0]
    pattern = re.compile(r'\s*(\d+(?:\.\d+)*)\s*<([^>]+)>')
    id_to_tag: Dict[str, str] = {}
    for i in range(header_idx + 1, len(df)):
        row = df.iloc[i]
        for j in range(len(df.columns)):
            val = row[j]
            if isinstance(val, str) and '<' in val:
                m = pattern.match(val)
                if m:
                    id_code, tag = m.groups()
                    id_to_tag[id_code] = tag
                    break
    return id_to_tag


def build_id_to_path(id_to_tag: Dict[str, str]) -> Dict[str, List[str]]:
    """Construct a mapping from ID code to the list of tag names along the path."""
    id_to_path: Dict[str, List[str]] = {}
    for id_code, tag in id_to_tag.items():
        parts = id_code.split('.')
        path: List[str] = []
        for i in range(1, len(parts) + 1):
            prefix = '.'.join(parts[:i])
            t = id_to_tag.get(prefix)
            if t:
                path.append(t)
        id_to_path[id_code] = path
    return id_to_path


def extract_leaf_ids(ws) -> Dict[int, str]:
    """
    Determine the leaf ID for each column by scanning from the bottom
    row upwards. The first cell containing an ID definition per
    column is considered the leaf.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    col_to_id: Dict[int, str] = {}
    pattern = re.compile(r'^(\d+(?:\.\d+)*)\s*<')
    for col in range(1, max_col + 1):
        id_code = None
        for row in range(max_row, 0, -1):
            value = ws.cell(row=row, column=col).value
            if value and isinstance(value, str) and '<' in value:
                m = pattern.match(value)
                if m:
                    id_code = m.group(1)
                    break
        if id_code:
            col_to_id[col] = id_code
    return col_to_id


def parse_xml_values(xml_file: str, id_to_path: Dict[str, List[str]], leaf_ids: Dict[int, str], max_col: int) -> List[str]:
    """Extract values for each column from the XML. Multi-valued tags are encoded as JSON arrays."""
    def ns_strip(tag: str) -> str:
        return tag.split('}')[-1] if '}' in tag else tag

    root = ET.parse(xml_file).getroot()
    row_values: List[str] = [''] * max_col
    for col, id_code in leaf_ids.items():
        path = id_to_path.get(id_code)
        if not path:
            continue
        nodes = [root]
        for tag in path:
            next_nodes = []
            for node in nodes:
                for child in node:
                    if ns_strip(child.tag) == tag:
                        next_nodes.append(child)
            nodes = next_nodes
            if not nodes:
                break
        vals: List[str] = []
        for node in nodes:
            if node.text:
                text = node.text.strip()
                if text:
                    vals.append(text)
        if not vals:
            row_values[col - 1] = ''
        elif len(vals) == 1:
            row_values[col - 1] = vals[0]
        else:
            row_values[col - 1] = json.dumps(vals, ensure_ascii=False)
    return row_values


def append_xml_to_excel(xml_file: str, representation: str, mapping_xlsx: str, excel_out: str) -> None:
    """Read a single XML and append its data row to the Excel file."""
    id_to_tag = build_id_to_tag_map(representation)
    id_to_path = build_id_to_path(id_to_tag)
    wb_map = openpyxl.load_workbook(mapping_xlsx)
    ws_map = wb_map.active
    leaf_ids = extract_leaf_ids(ws_map)
    max_col = ws_map.max_column
    # Parse values
    row = parse_xml_values(xml_file, id_to_path, leaf_ids, max_col)
    # Append to output workbook
    wb_out = openpyxl.load_workbook(excel_out)
    ws_out = wb_out.active
    ws_out.append(row)
    wb_out.save(excel_out)


def batch_process() -> None:
    """Process all XML files in the XML subfolder and append to the JSON Excel file."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    xml_dir = os.path.join(base_dir, 'XML')
    representation = os.path.join(base_dir, 'RappresentazioneTabellareFattOrdinaria (1).xlsx')
    mapping_xlsx = os.path.join(base_dir, 'hierarchical_mapping_fixed.xlsx')
    excel_out = os.path.join(base_dir, 'hierarchical_mapping_populated_json.xlsx')
    if not os.path.exists(xml_dir):
        print(f"XML directory '{xml_dir}' does not exist. No files processed.")
        return
    xml_files = [f for f in os.listdir(xml_dir) if f.lower().endswith('.xml')]
    if not xml_files:
        print(f"No XML files found in '{xml_dir}'.")
        return
    for xml_name in xml_files:
        xml_path = os.path.join(xml_dir, xml_name)
        print(f"Processing {xml_name}...")
        append_xml_to_excel(xml_path, representation, mapping_xlsx, excel_out)
    print(f"Finished processing {len(xml_files)} XML files. Data appended to {excel_out}.")


if __name__ == '__main__':
    batch_process()